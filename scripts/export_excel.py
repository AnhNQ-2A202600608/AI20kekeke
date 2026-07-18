import os

import openpyxl


def clean_val(val):
    if val is None:
        return ""
    return str(val).strip().replace("\n", " ").replace("|", "\\|")


def export_xlsx_to_md(xlsx_path, md_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    os.makedirs(os.path.dirname(md_path), exist_ok=True)

    with open(md_path, "w", encoding="utf-8") as f:
        f.write(f"# Excel Summary: {os.path.basename(xlsx_path)}\n\n")

        for name in wb.sheetnames:
            sheet = wb[name]
            f.write(f"## Sheet: {name}\n\n")

            # Read all rows
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                f.write("*Empty sheet*\n\n")
                continue

            # Filter out completely empty rows
            non_empty_rows = []
            for r in rows:
                if any(x is not None for x in r):
                    non_empty_rows.append(r)

            if not non_empty_rows:
                f.write("*Empty sheet*\n\n")
                continue

            # Let's find max column count
            max_cols = max(len(r) for r in non_empty_rows)

            # We will format as Markdown table
            # First row is headers
            headers = [clean_val(x) for x in non_empty_rows[0]]
            # Ensure headers are not all empty, if so use Col 1, Col 2, etc.
            if all(h == "" for h in headers):
                headers = [f"Col {i + 1}" for i in range(max_cols)]
            else:
                # pad headers if shorter
                while len(headers) < max_cols:
                    headers.append(f"Col {len(headers) + 1}")

            f.write("| " + " | ".join(headers) + " |\n")
            f.write("| " + " | ".join(["---"] * len(headers)) + " |\n")

            for row in non_empty_rows[1:]:
                vals = [clean_val(row[i]) if i < len(row) else "" for i in range(max_cols)]
                f.write("| " + " | ".join(vals) + " |\n")

            f.write("\n\n")

    print(f"Exported to {md_path}")


if __name__ == "__main__":
    xlsx = "team_project_management_template_ai_tutor_filled_with_progress_styled_summary.xlsx"
    md = "docs/research/excel_summary.md"
    export_xlsx_to_md(xlsx, md)
